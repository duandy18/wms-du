# app/tms/billing/importer.py
from __future__ import annotations

from datetime import datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO

from openpyxl import load_workbook

from .contracts import CarrierBillImportRowErrorData


def _to_clean_str(value: object) -> str | None:
    if value is None:
        return None
    s = str(value).strip()
    return s or None


def _to_decimal(value: object, *, field_label: str, row_no: int) -> Decimal | None:
    if value is None:
        return None
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)):
        return Decimal(str(value))

    s = str(value).strip()
    if s == "":
        return None

    try:
        return Decimal(s)
    except InvalidOperation as exc:
        raise ValueError(f"{field_label} 不是合法数字") from exc


def _to_datetime(value: object, *, field_label: str, row_no: int) -> datetime | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value

    s = str(value).strip()
    if s == "":
        return None

    candidates = [
        "%Y-%m-%d %H:%M:%S",
        "%Y/%m/%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%Y/%m/%d %H:%M",
        "%Y-%m-%d",
        "%Y/%m/%d",
    ]
    for fmt in candidates:
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue

    try:
        return datetime.fromisoformat(s)
    except ValueError as exc:
        raise ValueError(f"{field_label} 不是合法时间") from exc


def _is_blank_row(values: list[object | None]) -> bool:
    for v in values:
        if v is None:
            continue
        if isinstance(v, str) and v.strip() == "":
            continue
        return False
    return True


def parse_and_normalize_carrier_bill_xlsx(
    file_bytes: bytes,
) -> tuple[list[dict[str, object]], list[CarrierBillImportRowErrorData], int]:
    wb = load_workbook(
        filename=BytesIO(file_bytes),
        read_only=True,
        data_only=True,
    )
    ws = wb[wb.sheetnames[0]]

    iterator = ws.iter_rows(values_only=True)
    try:
        header_row = next(iterator)
    except StopIteration:
        return [], [CarrierBillImportRowErrorData(row_no=1, message="Excel 为空")], 0

    headers = [_to_clean_str(v) or f"__col_{idx+1}" for idx, v in enumerate(header_row)]

    valid_rows: list[dict[str, object]] = []
    errors: list[CarrierBillImportRowErrorData] = []
    skipped_count = 0

    for row_no, row in enumerate(iterator, start=2):
        values = list(row)
        if _is_blank_row(values):
            skipped_count += 1
            continue

        raw_payload: dict[str, object] = {
            headers[idx]: values[idx] if idx < len(values) else None
            for idx in range(len(headers))
        }

        try:
            tracking_no = _to_clean_str(raw_payload.get("运单号"))
            if not tracking_no:
                raise ValueError("运单号为空")

            business_time = _to_datetime(
                raw_payload.get("业务时间"),
                field_label="业务时间",
                row_no=row_no,
            )

            destination_province = _to_clean_str(raw_payload.get("目的省份"))
            destination_city = _to_clean_str(raw_payload.get("目的城市"))

            billing_weight = _to_decimal(
                raw_payload.get("结算重量"),
                field_label="结算重量",
                row_no=row_no,
            )
            freight_amount = _to_decimal(
                raw_payload.get("中转费/运费"),
                field_label="中转费/运费",
                row_no=row_no,
            )
            surcharge_amount = _to_decimal(
                raw_payload.get("附加费"),
                field_label="附加费",
                row_no=row_no,
            )

            freight_dec = freight_amount or Decimal("0")
            surcharge_dec = surcharge_amount or Decimal("0")
            total_amount = freight_dec + surcharge_dec

            valid_rows.append(
                {
                    "tracking_no": tracking_no,
                    "business_time": business_time,
                    "destination_province": destination_province,
                    "destination_city": destination_city,
                    "billing_weight_kg": float(billing_weight) if billing_weight is not None else None,
                    "freight_amount": float(freight_amount) if freight_amount is not None else None,
                    "surcharge_amount": float(surcharge_amount) if surcharge_amount is not None else None,
                    "total_amount": float(total_amount),
                    "settlement_object": _to_clean_str(raw_payload.get("结算对象")),
                    "order_customer": _to_clean_str(raw_payload.get("订单客户")),
                    "sender_name": _to_clean_str(raw_payload.get("寄件人")),
                    "network_name": _to_clean_str(raw_payload.get("所属网点")),
                    "size_text": _to_clean_str(raw_payload.get("长宽高")),
                    "parent_customer": _to_clean_str(raw_payload.get("父客户")),
                    "raw_payload": raw_payload,
                }
            )
        except ValueError as exc:
            errors.append(CarrierBillImportRowErrorData(row_no=row_no, message=str(exc)))

    return valid_rows, errors, skipped_count
